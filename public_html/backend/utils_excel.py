from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def set_value_safe(ws, cell, value):
    """Escribe en la celda superior izquierda si la celda está fusionada."""
    c = ws[cell]
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= c.row <= max_row and min_col <= c.column <= max_col:
            ws.cell(row=min_row, column=min_col, value=value)
            return
    ws[cell] = value


def _set_if_not_none(ws, cell, value):
    if value is not None:
        set_value_safe(ws, cell, value)


def escribir_datos_excel(path_excel, ciudad=None, zona=None, tipo_inst=None, mensual=None, anual=None, lat=None, lng=None):
    """Escribe solo los valores provistos (no sobrescribe None)."""
    wb = load_workbook(path_excel)
    ws = wb["Datos de Entrada"]
    _set_if_not_none(ws, "B7",  ciudad)     # Ciudad
    _set_if_not_none(ws, "B9",  zona)       # Zona
    _set_if_not_none(ws, "B11", tipo_inst)  # Tipo instalación
    _set_if_not_none(ws, "B12", mensual)    # kWh/mes
    _set_if_not_none(ws, "B13", anual)      # kWh/año
    _set_if_not_none(ws, "B15", lat)
    _set_if_not_none(ws, "B16", lng)
    wb.save(path_excel)
    wb.close()


def leer_resultados_excel(path_excel, min_row=3, max_row=50, min_col=2, max_col=12):
    wb = load_workbook(path_excel, data_only=True)
    ws = wb["Resultados"]
    datos = [
        list(row)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
    ]
    wb.close()
    return datos
