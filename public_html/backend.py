from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from pathlib import Path
import os

app = Flask(__name__)
CORS(app)

# === Ruta del Excel relativa a este archivo (backend.py) ===
# Se asume que el .xlsx está dentro de la carpeta "backend/" junto a este backend.py
BASE_DIR = Path(__file__).resolve().parent
EXCEL_DIR = BASE_DIR / 'backend'
# Nombre esperado; si cambia la versión, se intenta encontrar por patrón
EXCEL_PATH = EXCEL_DIR / 'Calculador Solar - web 06-24_con ayuda - modificaciones 2025_5.xlsx'
if not EXCEL_PATH.exists():
    matches = sorted(EXCEL_DIR.glob('Calculador Solar*.xlsx'))
    if matches:
        EXCEL_PATH = matches[0]
    else:
        raise FileNotFoundError(f'No se encontró el Excel dentro de: {EXCEL_DIR}')


@app.post('/api/guardar_ciudad')
def guardar_ciudad():
    """
    Guarda SOLO la ciudad en 'Datos de Entrada'!B7 y marca recálculo al abrir.
    No guarda lat/lng ni otras celdas.
    """
    data = request.get_json(silent=True) or {}
    ciudad = (data.get('ciudad') or '').strip()
    if not ciudad:
        return jsonify({'error': 'Falta el nombre de la ciudad'}), 400

    try:
        wb = load_workbook(EXCEL_PATH, data_only=False)
        if 'Datos de Entrada' not in wb.sheetnames:
            return jsonify({'error': "No existe la hoja 'Datos de Entrada'"}), 500
        ws = wb['Datos de Entrada']
        ws['B7'] = ciudad
        # Forzar recálculo cuando se abra en Excel/LibreOffice
        wb.calculation_properties.fullCalcOnLoad = True
        wb.save(EXCEL_PATH)
        return jsonify({'ok': True, 'ciudad': ciudad}), 200
    except Exception as e:
        return jsonify({'error': f'No se pudo escribir en Excel: {e}'}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
