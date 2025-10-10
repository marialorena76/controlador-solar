import openpyxl
import xlwings as xw
import os
import json
import platform

# Define the path to the Excel file
EXCEL_FILE = os.path.join('backend', 'Calculador Solar - web 06-24_con ayuda - modificaciones 2025_5.xlsx')

# --- 1. Define Input and Output Mappings ---

# Map user_data keys to specific cells in the "Datos de Entrada" sheet
INPUT_CELL_MAP = {
    "city": "B3",
    "user_level": "B4",
}

# Map titles to the result cells to be read after recalculation
OUTPUT_CELL_MAP = {
    "Datos técnicos del dimensionamiento": {
        "Consumo anual de energía eléctrica": "C6",
        "Generación anual de energía eléctrica": "C7",
        "Energía para autoconsumo": "C8",
        "Energía inyectada a la red": "C9",
        "Potencia de paneles sugerida": "C12",
        "Cantidad de paneles necesarios": "C13",
        "Superficie necesaria": "C14",
        "Vida útil del proyecto (años)": "C16"
    },
    "Resultados económicos": {
        "Saldo neto anual a favor si realiza instalación": "C21",
        "Gasto anual si NO instala": "C23",
        "Inversión inicial necesaria": "C25"
    },
    "Contribución a la mitigación del cambio climático": {
        "Emisiones de GEI evitadas": "C29"
    }
}


def write_to_excel(user_data):
    """
    Writes user data to the specified Excel file using openpyxl.
    """
    print(f"Writing data to {EXCEL_FILE}...")
    try:
        # Load the workbook with data_only=False to preserve formulas
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws = wb["Datos de Entrada"]

        # Write each value from user_data to its mapped cell
        for key, cell in INPUT_CELL_MAP.items():
            if key in user_data:
                ws[cell] = user_data[key]
                print(f"  - Wrote '{user_data[key]}' to cell {cell}")

        wb.save(EXCEL_FILE)
        print("Successfully saved the workbook with new inputs.")
        return True
    except FileNotFoundError:
        print(f"ERROR: Excel file not found at {EXCEL_FILE}")
        return False
    except KeyError:
        print("ERROR: Sheet 'Datos de Entrada' not found in the workbook.")
        return False
    except Exception as e:
        print(f"An unexpected error occurred during writing: {e}")
        return False


def recalculate_excel():
    """
    Forces recalculation of the Excel workbook using xlwings.
    Includes fallback for non-Windows environments.
    """
    print("\nAttempting to recalculate Excel formulas...")
    if platform.system() != "Windows":
        print("WARNING: xlwings is not running on Windows. Recalculation depends on a running Excel instance, which is unlikely to be available.")
        print("The script will proceed, but the read values might not be updated.")
        # Future enhancement: Implement LibreOffice headless mode here if needed.
        return

    app = None
    try:
        with xw.App(visible=False, add_book=False) as app:
            wb_x = app.books.open(EXCEL_FILE)
            print("  - Workbook opened with xlwings.")
            wb_x.app.calculate()
            print("  - Recalculation triggered.")
            wb_x.save()
            print("  - Workbook saved after calculation.")
            wb_x.close()
            print("  - Workbook closed.")
        print("Excel recalculation successful.")
    except Exception as e:
        print(f"ERROR: Failed to recalculate with xlwings. Reason: {e}")
        print("This may be because Excel is not installed or the file is corrupted/locked.")
        print("Proceeding to read data, but it may not be up-to-date.")
    finally:
        if app and app.pid:
            app.quit()


def read_from_excel():
    """
    Reads the calculated values from the Excel file using openpyxl.
    """
    print(f"\nReading calculated data from {EXCEL_FILE}...")
    results = {}
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # From inspection, the correct sheet name is "Resultados".
        sheet_name = "Resultados"
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: The expected results sheet '{sheet_name}' was not found in the workbook.")
            print(f"Available sheets are: {wb.sheetnames}")
            return None
        
        ws = wb[sheet_name]
        print(f"  - Reading from sheet: '{sheet_name}'")

        for section, items in OUTPUT_CELL_MAP.items():
            results[section] = {}
            for title, cell_ref in items.items():
                cell = ws[cell_ref]
                value = cell.value
                
                # Get unit from the adjacent column (D)
                unit_cell_ref = f'D{cell.row}'
                unit = ws[unit_cell_ref].value
                
                if value is None:
                    print(f"  - WARNING: Value for '{title}' in cell {cell_ref} is None (not calculated).")
                    results[section][title] = "Valor no calculado"
                else:
                    if unit and isinstance(unit, str) and unit.strip():
                         results[section][title] = f"{value} {unit.strip()}"
                    else:
                         results[section][title] = value
                    
                    print(f"  - Read '{title}': {results[section][title]}")

        return results

    except FileNotFoundError:
        print(f"ERROR: Excel file not found at {EXCEL_FILE}")
        return None
    except KeyError as e:
        print(f"ERROR: A required sheet was not found. Details: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred during reading: {e}")
        return None


def generate_report(results):
    """
    Generates a text report and a JSON output from the results dictionary.
    """
    if not results:
        print("\nNo results to generate a report.")
        return

    print("\n--- Informe de Cálculo Solar ---")
    text_report_lines = []
    for section, items in results.items():
        print(f"\n## {section}")
        text_report_lines.append(f"## {section}")
        for title, value in items.items():
            report_line = f"- {title}: {value}"
            print(report_line)
            text_report_lines.append(report_line)

    # Brief conclusion
    try:
        saldo_anual = results.get("Resultados económicos", {}).get("Saldo neto anual a favor si realiza instalación", "N/A")
        inversion = results.get("Resultados económicos", {}).get("Inversión inicial necesaria", "N/A")
        emisiones = results.get("Contribución a la mitigación del cambio climático", {}).get("Emisiones de GEI evitadas", "N/A")

        conclusion = f"""
## Conclusión Breve
El proyecto presenta un saldo neto anual favorable de {saldo_anual} frente a una inversión inicial de {inversion}.
Desde el punto de vista ambiental, se estima que se evitarán {emisiones} de emisiones de gases de efecto invernadero durante la vida útil del proyecto.
"""
        print(conclusion)
        text_report_lines.append(conclusion)
    except Exception as e:
        print(f"\nCould not generate conclusion due to missing data: {e}")

    print("\n--- JSON Output ---")
    print(json.dumps(results, indent=2, ensure_ascii=False))


def main():
    """
    Main function to orchestrate the report generation process.
    """
    print("Starting report generation process...")

    # Example user data
    user_data = {
        "city": "Córdoba",
        "user_level": "Residencial"
    }
    
    if not os.path.exists(EXCEL_FILE):
        print(f"CRITICAL: The Excel file is missing at the expected path: {EXCEL_FILE}")
        return

    # 1. Write user data to Excel
    if not write_to_excel(user_data):
        return

    # 2. Force recalculation
    recalculate_excel()

    # 3. Read calculated values
    final_results = read_from_excel()

    # 4. Generate and display the report
    generate_report(final_results)


if __name__ == "__main__":
    main()