# backend.py
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import pandas as pd
import os
import json
import math
from threading import Lock
from openpyxl import load_workbook
#from . import engine

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCRIPT_DIR = BASE_DIR
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

DEFAULT_EXCEL_FILENAME = 'Calculador Solar - web 06-24_con ayuda - modificaciones 2025_4.xlsx'
EXCEL_RELATIVE_DIR = 'backend'
EXCEL_FILE_PATH = os.path.join(SCRIPT_DIR, EXCEL_RELATIVE_DIR, DEFAULT_EXCEL_FILENAME)
EXCEL_PATH = os.path.abspath(EXCEL_FILE_PATH)

excel_lock = Lock()
# Nota: este candado se reutiliza tanto para la escritura directa en el archivo
# Excel como para las ejecuciones del motor de cálculo, de modo que todas las
# operaciones que interactúan con el libro queden coordinadas y se evite el
# acceso concurrente.

# The write_to_debug_log function is being removed and replaced with standard prints.

def clean_nan_in_data(obj):
    """
    Recursively walk a dict or list and replace float('nan') with None,
    as NaN is not valid in JSON.
    """
    if isinstance(obj, dict):
        return {k: clean_nan_in_data(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [clean_nan_in_data(elem) for elem in obj]
    # Check specifically for float('nan')
    if isinstance(obj, float) and math.isnan(obj):
        return None
    return obj

CONSUMOS_JSON_PATH = os.path.join(PROJECT_ROOT, 'consumos_electrodomesticos.json')


app = Flask(__name__, static_folder=PROJECT_ROOT, static_url_path='')
CORS(app)  # Habilita CORS para permitir solicitudes desde el frontend


# --- Ruta para actualizar una celda específica del Excel ---
@app.route('/api/excel/update_cell', methods=['POST'])
def update_excel_cell():
    try:
        payload = request.get_json(silent=True) or {}
        raw_city_name = payload.get('ciudad', '')
        # Manejar valores None o que no sean strings convirtiéndolos explícitamente
        city_name = '' if raw_city_name is None else str(raw_city_name)

        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({"success": False, "error": "Archivo Excel de configuración no encontrado."}), 404

        with excel_lock:
            workbook = None
            try:
                workbook = load_workbook(EXCEL_FILE_PATH)
                worksheet = workbook['Datos de Entrada']
                worksheet['B7'] = city_name
                workbook.save(EXCEL_FILE_PATH)
            except KeyError:
                return jsonify({"success": False, "error": "Hoja 'Datos de Entrada' no encontrada en el Excel."}), 500
            except (OSError, IOError) as save_error:
                return jsonify({"success": False, "error": f"No se pudo guardar el archivo Excel: {save_error}"}), 500
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        # Ignoramos el error de cierre para no sobreescribir la respuesta original.
                        pass

        return jsonify({"success": True})

    except (OSError, IOError) as io_error:
        return jsonify({"success": False, "error": f"Error de E/S al actualizar el Excel: {io_error}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/excel/update_cell: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "error": f"Error interno del servidor: {e}"}), 500


# --- NUEVA RUTA: Para obtener la lista de ciudades ---
@app.route('/api/ciudades', methods=['GET'])
def obtener_ciudades():
    try:
        if not os.path.exists(EXCEL_PATH):
            return jsonify({"error": f"Archivo Excel no encontrado en {EXCEL_PATH}"}), 404

        try:
            df_ciudades = pd.read_excel(EXCEL_PATH, sheet_name='Ciudades', engine='openpyxl')
        except ValueError:
            return jsonify({"error": "Hoja de ciudades no encontrada. TODO: ajustar nombre de la hoja que contiene la lista de ciudades."}), 500

        columna_ciudad = 'Ciudad'
        if columna_ciudad not in df_ciudades.columns:
            return jsonify({"error": "Columna 'Ciudad' no encontrada. TODO: actualizar el nombre de la columna con los nombres de ciudad."}), 500

        ciudades = [str(c).strip() for c in df_ciudades[columna_ciudad].dropna() if str(c).strip()]
        return jsonify({"ciudades": ciudades})

    except Exception as exc:  # noqa: BLE001
        import traceback
        print(f"ERROR en /api/ciudades: {exc}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno al obtener ciudades: {exc}"}), 500


# --- NUEVA RUTA: Para guardar la ciudad seleccionada ---
@app.route('/api/guardar_ciudad', methods=['POST'])
def guardar_ciudad():
    payload = request.get_json(silent=True) or {}
    ciudad = (payload.get('ciudad') or payload.get('city') or '').strip()

    if not ciudad:
        return jsonify({"ok": False, "error": "Falta el campo 'ciudad' en la solicitud."}), 400

    if not os.path.exists(EXCEL_PATH):
        return jsonify({"ok": False, "error": f"Archivo Excel no encontrado en {EXCEL_PATH}"}), 404

    with excel_lock:
        wb = None
        try:
            wb = load_workbook(EXCEL_PATH, data_only=False)
            if 'Datos de Entrada' not in wb.sheetnames:
                return jsonify({"ok": False, "error": "Hoja 'Datos de Entrada' no encontrada en el Excel."}), 500

            ws = wb['Datos de Entrada']
            ws['B7'] = ciudad  # La celda B7 se usa actualmente para la ciudad de instalación
            wb.save(EXCEL_PATH)
        except Exception as exc:  # noqa: BLE001
            import traceback
            print(f"ERROR en /api/guardar_ciudad: {exc}")
            print(traceback.format_exc())
            return jsonify({"ok": False, "error": f"No se pudo guardar la ciudad: {exc}"}), 500
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    return jsonify({"ok": True, "ciudad": ciudad})


# --- NUEVA RUTA: Para obtener la lista de electrodomésticos y sus consumos ---
@app.route('/api/electrodomesticos', methods=['GET'])
def get_electrodomesticos_consumos():
    """
    Lee los datos de electrodomésticos desde consumos_electrodomesticos.json,
    calcula el consumo diario en kWh y los devuelve en el formato esperado.
    """
    try:
        print(f"DEBUG: Solicitud a /api/electrodomesticos. Leyendo desde: {CONSUMOS_JSON_PATH}")
        if not os.path.exists(CONSUMOS_JSON_PATH):
            print(f"ERROR CRITICO: El archivo {CONSUMOS_JSON_PATH} no fue encontrado.")
            return jsonify({"error": f"Archivo de datos de electrodomésticos no encontrado en el servidor."}), 404

        with open(CONSUMOS_JSON_PATH, 'r', encoding='utf-8') as f:
            source_data = json.load(f)

        # El frontend espera una estructura {"categorias": {"cat1": [item1, ...], "cat2": [...]}}
        # El archivo JSON ya tiene la estructura de categorías, solo necesitamos procesar los items.

        categorized_appliances = {}
        for category, items in source_data.items():
            if category not in categorized_appliances:
                categorized_appliances[category] = []

            for item in items:
                watts = item.get('watts', 0)
                hours = item.get('hoursPerDay', 0)

                # Calcular consumo diario en kWh
                consumo_diario_kwh = (watts * hours) / 1000.0

                appliance_entry = {
                    "name": item.get('name', 'Sin Nombre'),
                    "consumo_diario_kwh": consumo_diario_kwh,
                    "watts": watts
                }
                categorized_appliances[category].append(appliance_entry)

        print(f"DEBUG: Se procesaron {sum(len(v) for v in categorized_appliances.values())} electrodomésticos desde JSON.")
        return jsonify({"categorias": categorized_appliances})

    except json.JSONDecodeError as je:
        print(f"ERROR en /api/electrodomesticos: Error al decodificar {CONSUMOS_JSON_PATH}: {je}")
        return jsonify({"error": "El archivo de datos de electrodomesticos está corrupto."}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/electrodomesticos: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

# --- Ruta para generar informe (actualizada) ---
@app.route('/api/generar_informe', methods=['POST'])
def generar_informe():
    try:
        payload = request.get_json(force=True) or {}
        app.logger.info(f"Payload recibido: {payload}")

        if not os.path.exists(EXCEL_PATH):
            return jsonify({"error": f"No se encontró el archivo Excel en {EXCEL_PATH}"}), 500

        ciudad = (payload.get("ciudad") or payload.get("city") or "").strip()
        zona = (payload.get("zonaInstalacionBasic") or "").strip()
        total_mes = float(payload.get("totalMonthlyConsumption", 0) or 0)
        total_anio = float(payload.get("totalAnnualConsumption", 0) or 0)

        if total_anio <= 0:
            return jsonify({"error": "Consumo anual inválido (0 o negativo)"}), 400

        # Abrir el Excel y escribir datos en la hoja de entrada
        wb = load_workbook(EXCEL_PATH)
        if "Datos de Entrada" not in wb.sheetnames:
            wb.close()
            return jsonify({"error": "No se encontró la hoja 'Datos de Entrada'"}), 500
        ws_in = wb["Datos de Entrada"]

        try:
            if ciudad:
                ws_in["B7"] = ciudad
            if zona:
                ws_in["B9"] = zona
            ws_in["B11"] = total_mes
            ws_in["B12"] = total_anio
        except Exception as e:
            wb.close()
            app.logger.exception("Error escribiendo celdas")
            return jsonify({"error": f"Error al escribir celdas: {str(e)}"}), 500

        wb.save(EXCEL_PATH)
        wb.close()

        # Reabrir con data_only=True para leer fórmulas calculadas
        wb2 = load_workbook(EXCEL_PATH, data_only=True)
        if "Resultados" not in wb2.sheetnames:
            wb2.close()
            return jsonify({"error": "No se encontró la hoja 'Resultados'"}), 500

        ws_out = wb2["Resultados"]
        datos = []
        for row in ws_out.iter_rows(min_row=3, max_row=50, min_col=2, max_col=12, values_only=True):
            datos.append(list(row))
        wb2.close()

        app.logger.info("Informe generado correctamente.")
        return jsonify({
            "status": "ok",
            "resumen": {
                "ciudad": ciudad or None,
                "zona": zona or None,
                "consumo_mensual_kwh": total_mes,
                "consumo_anual_kwh": total_anio
            },
            "tabla_resultados": datos
        }), 200

    except Exception as e:
        app.logger.exception("Excepción en generar_informe")
        return jsonify({"error": f"Excepción en generar_informe: {str(e)}"}), 500

# Opcional: Rutas para servir los archivos estáticos de tu frontend
@app.route('/')
def serve_index_html():
    return send_from_directory(PROJECT_ROOT, 'index.html')

@app.route('/<path:path>')
def serve_static_files(path):
    return send_from_directory(PROJECT_ROOT, path)


# --- NUEVA RUTA: Para obtener opciones de superficie de instalación ---
@app.route('/api/superficie_options', methods=['GET'])
def get_superficie_options():
    try:
        print(f"DEBUG: Solicitud a /api/superficie_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        # Usamos decimal=',' por si los números en la hoja 'Tablas' usan coma decimal.
        # Si esta parte especfica de la hoja usa punto decimal, se puede omitir.
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de superficie.")

        # Column M (descripción) es índice 12, Column N (valor) es índice 13
        col_desc_idx = 12
        col_valor_idx = 13
        # Filas 58 a 79 en Excel son iloc 57 a 78 (porque iloc es [start, end-1] para end)
        # Para incluir la fila 79 (índice 78), el rango de iloc debe ir hasta 79.
        fila_inicio_idx = 57 # Fila 58 en Excel
        fila_fin_idx = 78    # Fila 79 en Excel

        superficie_options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_desc_idx >= max_cols_df or col_valor_idx >= max_cols_df:
            print(f"WARN: Columnas M ({col_desc_idx}) o N ({col_valor_idx}) fuera de límites (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definicin de columnas fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1): # +1 para incluir fila_fin_idx
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura de opciones de superficie detenida.")
                break

            descripcion = df_tablas.iloc[r_idx, col_desc_idx]
            valor = df_tablas.iloc[r_idx, col_valor_idx]

            if pd.isna(descripcion) or str(descripcion).strip() == "":
                print(f"DEBUG: Fila {r_idx+1} omitida por descripción NaN o vacía para opciones de superficie.")
                continue

            valor_float = 0.0
            if pd.isna(valor):
                print(f"DEBUG: Valor NaN para superficie '{descripcion}' en fila {r_idx+1}, usando 0.0.")
            else:
                try:
                    # Si decimal=',' ya manejó la coma, esto debería ser directo.
                    # Si no, y los números pueden tener comas, necesitaríamos: str(valor).replace(',', '.')
                    valor_float = float(valor)
                except ValueError:
                    print(f"WARN: No se pudo convertir valor de superficie '{valor}' a float para '{descripcion}'. Usando 0.0.")

            superficie_options_lista.append({
                "descripcion": str(descripcion),
                "valor": valor_float
            })

        print(f"DEBUG: Total opciones de superficie leídas de 'Tablas' M{fila_inicio_idx+1}:N{fila_fin_idx+1}: {len(superficie_options_lista)}")
        return jsonify(superficie_options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/superficie_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        # Esto podría ocurrir si 'Tablas' no existe, o si las columnas M/N no existen (aunque ya chequeamos índices).
        print(f"ERROR en /api/superficie_options: Hoja 'Tablas' o columnas M/N no encontradas? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo (¿nombre de hoja o columna incorrecto?): {e}"}), 500
    except IndexError as e:
        # Esto podría ocurrir si el rango de filas/columnas es incorrecto a pesar de las verificaciones.
        print(f"ERROR en /api/superficie_options: Rango de celdas fuera de límites: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/superficie_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de superficie: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener opciones de rugosidad ---
@app.route('/api/rugosidad_options', methods=['GET'])
def get_rugosidad_options():
    try:
        print(f"DEBUG: Solicitud a /api/rugosidad_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leda para opciones de rugosidad.")

        col_desc_idx = 12  # Columna M
        col_valor_idx = 13 # Columna N
        # Filas Excel 82 a 86 -> iloc 81 a 85
        fila_inicio_idx = 81 # Fila 82 en Excel
        fila_fin_idx = 85    # Fila 86 en Excel

        rugosidad_options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        # Validar que las columnas existan en el DataFrame
        if col_desc_idx >= max_cols_df or col_valor_idx >= max_cols_df:
            print(f"WARN: Columnas M ({col_desc_idx}) o N ({col_valor_idx}) fuera de límites (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columnas para rugosidad fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1): # +1 para incluir fila_fin_idx
            # Validar que la fila exista
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para rugosidad fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            descripcion = df_tablas.iloc[r_idx, col_desc_idx]
            valor = df_tablas.iloc[r_idx, col_valor_idx]

            if pd.isna(descripcion) or str(descripcion).strip() == "":
                print(f"DEBUG: Fila {r_idx+1} omitida por descripción NaN o vacía para opciones de rugosidad.")
                continue

            valor_float = 0.0
            if pd.isna(valor):
                print(f"DEBUG: Valor NaN para rugosidad '{descripcion}' en fila {r_idx+1}, usando 0.0.")
            else:
                try:
                    valor_float = float(valor)
                except ValueError:
                    print(f"WARN: No se pudo convertir valor de rugosidad '{valor}' a float para '{descripcion}'. Usando 0.0.")

            rugosidad_options_lista.append({
                "descripcion": str(descripcion),
                "valor": valor_float
            })

        print(f"DEBUG: Total opciones de rugosidad leídas de 'Tablas' M{fila_inicio_idx+1}:N{fila_fin_idx+1}: {len(rugosidad_options_lista)}")
        return jsonify(rugosidad_options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/rugosidad_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/rugosidad_options: Hoja 'Tablas' o columnas M/N no encontradas? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para rugosidad (¿nombre de hoja o columna incorrecto?): {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/rugosidad_options: Rango de celdas fuera de límites para rugosidad: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites para rugosidad: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/rugosidad_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de rugosidad: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener opciones de rotación ---
@app.route('/api/rotacion_options', methods=['GET'])
def get_rotacion_options():
    try:
        print(f"DEBUG: Solicitud a /api/rotacion_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de rotación.")

        col_desc_idx = 15  # Columna P
        col_valor_idx = 16 # Columna Q
        # Filas Excel 96 a 103 -> iloc 95 a 102
        fila_inicio_idx = 95 # Fila 96 en Excel
        fila_fin_idx = 102   # Fila 103 en Excel

        rotacion_options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_desc_idx >= max_cols_df or col_valor_idx >= max_cols_df:
            print(f"WARN: Columnas P ({col_desc_idx}) o Q ({col_valor_idx}) fuera de límites (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columnas para rotación fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1): # +1 para incluir fila_fin_idx
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para rotación fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            descripcion = df_tablas.iloc[r_idx, col_desc_idx]
            valor = df_tablas.iloc[r_idx, col_valor_idx]

            if pd.isna(descripcion) or str(descripcion).strip() == "":
                print(f"DEBUG: Fila {r_idx+1} omitida por descripción NaN o vacía para opciones de rotación.")
                continue

            valor_float = 0.0
            if pd.isna(valor):
                print(f"DEBUG: Valor NaN para rotación '{descripcion}' en fila {r_idx+1}, usando 0.0.")
            else:
                try:
                    valor_float = float(valor)
                except ValueError:
                    print(f"WARN: No se pudo convertir valor de rotación '{valor}' a float para '{descripcion}'. Usando 0.0.")

            rotacion_options_lista.append({
                "descripcion": str(descripcion),
                "valor": valor_float
            })

        print(f"DEBUG: Total opciones de rotación leídas de 'Tablas' P{fila_inicio_idx+1}:Q{fila_fin_idx+1}: {len(rotacion_options_lista)}")
        return jsonify(rotacion_options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/rotacion_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/rotacion_options: Hoja 'Tablas' o columnas P/Q no encontradas? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para rotación (¿nombre de hoja o columna incorrecto?): {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/rotacion_options: Rango de celdas fuera de límites para rotación: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites para rotación: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/rotacion_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de rotación: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener opciones de Método de Cálculo ---
@app.route('/api/metodo_calculo_options', methods=['GET'])
def get_metodo_calculo_options():
    try:
        print(f"DEBUG: Solicitud a /api/metodo_calculo_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        # decimal=',' might not be strictly necessary if column I is purely text, but harmless.
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de método de cálculo.")

        col_idx = 8  # Columna I
        # Las filas exactas pueden variar según la versión del Excel.
        # Se lee un rango amplio y se filtran solo las entradas que
        # correspondan a los métodos de clculo deseados.
        fila_inicio_idx = 17  # Primer valor posible (Fila 18 en Excel)
        fila_fin_idx = 22     # Última fila revisada (Fila 23 en Excel)

        metodo_options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_idx >= max_cols_df:
            print(f"WARN: Columna I ({col_idx}) fuera de lmites (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columna para método de cálculo fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1):  # +1 para incluir fila_fin_idx
            if r_idx >= max_filas_df:
                print(
                    f"WARN: Fila {r_idx+1} para método de cálculo fuera de límites "
                    f"(hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida."
                )
                break

            valor_celda = df_tablas.iloc[r_idx, col_idx]

            if pd.isna(valor_celda) or str(valor_celda).strip() == "":
                print(
                    f"DEBUG: Fila {r_idx+1}, Columna I omitida por valor NaN o vacío para método de cálculo."
                )
                continue

            valor_str = str(valor_celda).strip()

            # Solo conservar valores que empiecen con "Cielo" para evitar
            # que se cuelen modelos como 'Modelo Liu-Jordan' en esta lista.
            if valor_str.startswith("Cielo"):
                metodo_options_lista.append(valor_str)

        print(f"DEBUG: Total opciones de método de cálculo leídas de 'Tablas' I{fila_inicio_idx+1}:I{fila_fin_idx+1}: {len(metodo_options_lista)}")
        return jsonify(metodo_options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/metodo_calculo_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/metodo_calculo_options: Hoja 'Tablas' o columna I no encontrada? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para método de cálculo (¿nombre de hoja o columna incorrecto?): {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/metodo_calculo_options: Rango de celdas fuera de límites para método de cálculo: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites para método de cálculo: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/metodo_calculo_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de método de cálculo: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener opciones de Modelo del Método ---
@app.route('/api/modelo_metodo_options', methods=['GET'])
def get_modelo_metodo_options():
    try:
        print(f"DEBUG: Solicitud a /api/modelo_metodo_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de modelo del método.")

        col_idx = 8  # Columna I
        # Filas Excel 20 a 23 -> iloc 19 a 22
        fila_inicio_idx = 19 # Fila 20 en Excel
        fila_fin_idx = 22    # Fila 23 en Excel

        modelo_options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_idx >= max_cols_df:
            print(f"WARN: Columna I ({col_idx}) fuera de límites para modelo del método (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definicin de columna para modelo del mtodo fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1): # +1 para incluir fila_fin_idx
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para modelo del mtodo fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            valor_celda = df_tablas.iloc[r_idx, col_idx]

            if pd.isna(valor_celda) or str(valor_celda).strip() == "":
                print(f"DEBUG: Fila {r_idx+1}, Columna I omitida por valor NaN o vacío para modelo del mtodo.")
                continue

            modelo_options_lista.append(str(valor_celda).strip())

        print(f"DEBUG: Total opciones de modelo del método leídas de 'Tablas' I{fila_inicio_idx+1}:I{fila_fin_idx+1}: {len(modelo_options_lista)}")
        return jsonify(modelo_options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/modelo_metodo_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/modelo_metodo_options: Hoja 'Tablas' o columna I no encontrada? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de clculo para modelo del método: {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/modelo_metodo_options: Rango de celdas fuera de límites para modelo del método: {e}")
        return jsonify({"error": f"Error de ndice, rango de celdas fuera de límites para modelo del método: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/modelo_metodo_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de modelo del método: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener opciones de Marca Panel ---
@app.route('/api/marca_panel_options', methods=['GET'])
def get_marca_panel_options():
    try:
        print(f"DEBUG: Solicitud a /api/marca_panel_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de marca de panel.")

        col_idx = 8  # Columna I
        # Filas Excel 38 a 42 -> iloc 37 a 41
        fila_inicio_idx = 37 # Fila 38 en Excel
        fila_fin_idx = 41    # Fila 42 en Excel

        options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_idx >= max_cols_df:
            print(f"WARN: Columna I ({col_idx}) fuera de límites para marca panel (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columna para marca panel fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1):
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para marca panel fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            valor_celda = df_tablas.iloc[r_idx, col_idx]

            if pd.isna(valor_celda) or str(valor_celda).strip() == "":
                print(f"DEBUG: Fila {r_idx+1}, Columna I omitida por valor NaN o vacío para marca panel.")
                continue

            options_lista.append(str(valor_celda).strip())

        print(f"DEBUG: Total opciones de marca panel leídas de 'Tablas' I{fila_inicio_idx+1}:I{fila_fin_idx+1}: {len(options_lista)}")
        return jsonify(options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/marca_panel_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/marca_panel_options: Hoja 'Tablas' o columna I no encontrada? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para marca panel: {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/marca_panel_options: Rango de celdas fuera de límites para marca panel: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de lmites para marca panel: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/marca_panel_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de marca panel: {str(e)}"}), 500

# --- NUEVA RUTA: Para obtener opciones de Modelo Temperatura Panel ---
@app.route('/api/modelo_temperatura_panel_options', methods=['GET'])
def get_modelo_temperatura_panel_options():
    try:
        print(f"DEBUG: Solicitud a /api/modelo_temperatura_panel_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de modelo temperatura panel.")

        col_idx = 8  # Columna I
        # Filas Excel 29 a 33 -> iloc 28 a 32
        fila_inicio_idx = 28 # Fila 29 en Excel
        fila_fin_idx = 32    # Fila 33 en Excel

        options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_idx >= max_cols_df:
            print(f"WARN: Columna I ({col_idx}) fuera de límites para modelo temperatura panel (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columna para modelo temperatura panel fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1):
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para modelo temperatura panel fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            valor_celda = df_tablas.iloc[r_idx, col_idx]

            if pd.isna(valor_celda) or str(valor_celda).strip() == "":
                print(f"DEBUG: Fila {r_idx+1}, Columna I omitida por valor NaN o vacío para modelo temperatura panel.")
                continue

            options_lista.append(str(valor_celda).strip())

        print(f"DEBUG: Total opciones de modelo temperatura panel leídas de 'Tablas' I{fila_inicio_idx+1}:I{fila_fin_idx+1}: {len(options_lista)}")
        return jsonify(options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/modelo_temperatura_panel_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/modelo_temperatura_panel_options: Hoja 'Tablas' o columna I no encontrada? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para modelo temperatura panel: {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/modelo_temperatura_panel_options: Rango de celdas fuera de lmites para modelo temperatura panel: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites para modelo temperatura panel: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/modelo_temperatura_panel_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de modelo temperatura panel: {str(e)}"}), 500


# --- NUEVA RUTA: Valor por defecto del Modelo de Temperatura de Paneles ---
@app.route('/api/modelo_temperatura_panel_valor', methods=['GET'])
def get_modelo_temperatura_panel_valor():
    """Obtener el valor de la celda C83 de 'Datos de Entrada'."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        import re

        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        def read_sheet(zf, path):
            with zf.open(path) as f:
                return ET.fromstring(f.read())

        def parse_shared_strings(zf):
            root = read_sheet(zf, "xl/sharedStrings.xml")
            return ["".join(t.text or "" for t in si.findall(".//s:t", ns)) for si in root.findall("s:si", ns)]

        def cell_value(root, ref, strings):
            cell = root.find(f".//s:c[@r='{ref}']", ns)
            if cell is None:
                return None
            v = cell.find("s:v", ns)
            if v is None:
                return ""
            return strings[int(v.text)] if cell.get("t") == "s" else v.text

        with zipfile.ZipFile(EXCEL_FILE_PATH) as zf:
            strings = parse_shared_strings(zf)
            datos = read_sheet(zf, "xl/worksheets/sheet3.xml")  # Datos de Entrada
            paneles = read_sheet(zf, "xl/worksheets/sheet14.xml")  # Paneles comerciales
            tablas = read_sheet(zf, "xl/worksheets/sheet12.xml")  # Tablas

            marca = cell_value(datos, "C81", strings)
            potencia = cell_value(datos, "C82", strings)
            clave = f"{marca}{potencia}"
            marca_generica = cell_value(tablas, "I42", strings)

            # Construir diccionario de busqueda A2:B135
            cells = {c.attrib["r"]: cell_value(paneles, c.attrib["r"], strings) for c in paneles.findall(".//s:c", ns)}
            modelo = None
            for fila in range(2, 136):
                if cells.get(f"A{fila}") == clave:
                    modelo = cells.get(f"B{fila}")
                    break

            if modelo is not None:
                valor = modelo
            else:
                valor = "" if marca == marca_generica else "VERIFIQUE LA POTENCIA SELECCIONADA, NO HAY PANELES DE ESA POTENCIA DE LA MARCA ELEGIDA"

        return jsonify({"valor": valor})

    except Exception as exc:  # noqa: BLE001
        import traceback
        print(f"ERROR en /api/modelo_temperatura_panel_valor: {exc}")
        print(traceback.format_exc())
        return jsonify({"error": "No se pudo obtener el valor solicitado."}), 500


# --- NUEVA RUTA: Para obtener opciones de Frecuencia de Lluvias ---
@app.route('/api/frecuencia_lluvias_options', methods=['GET'])
def get_frecuencia_lluvias_options():
    try:
        print(f"DEBUG: Solicitud a /api/frecuencia_lluvias_options. Leyendo de HOJA 'Tablas' desde: {EXCEL_FILE_PATH}")
        df_tablas = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Tablas', engine='openpyxl')
        print("DEBUG: Hoja 'Tablas' leída para opciones de frecuencia de lluvias.")

        col_data_idx = 12  # Column M
        fila_inicio_idx = 91 # Excel row 92 (92-1)
        fila_fin_idx = 94    # Excel row 95 (95-1)

        options_lista = []
        max_filas_df = df_tablas.shape[0]
        max_cols_df = df_tablas.shape[1]

        if col_data_idx >= max_cols_df:
            print(f"WARN: Columna M ({col_data_idx}) fuera de límites para frecuencia lluvias (hoja 'Tablas' tiene {max_cols_df} columnas).")
            return jsonify({"error": "Definición de columna para frecuencia lluvias fuera de los límites de la hoja."}), 500

        for r_idx in range(fila_inicio_idx, fila_fin_idx + 1):
            if r_idx >= max_filas_df:
                print(f"WARN: Fila {r_idx+1} para frecuencia lluvias fuera de límites (hoja 'Tablas' tiene {max_filas_df} filas). Lectura detenida.")
                break

            data_val = df_tablas.iloc[r_idx, col_data_idx]

            if pd.isna(data_val) or str(data_val).strip() == "":
                print(f"DEBUG: Fila {r_idx+1}, Columna M omitida por valor NaN o vacío para frecuencia lluvias.")
                continue

            options_lista.append(str(data_val).strip())

        print(f"DEBUG: Total opciones de frecuencia lluvias leídas de 'Tablas' M{fila_inicio_idx+1}:M{fila_fin_idx+1}: {len(options_lista)}")
        return jsonify(options_lista)

    except FileNotFoundError:
        print(f"ERROR en /api/frecuencia_lluvias_options: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError as e:
        print(f"ERROR en /api/frecuencia_lluvias_options: Hoja 'Tablas' o columna M no encontrada? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para frecuencia lluvias: {e}"}), 500
    except IndexError as e:
        print(f"ERROR en /api/frecuencia_lluvias_options: Rango de celdas fuera de límites para frecuencia lluvias: {e}")
        return jsonify({"error": f"Error de índice, rango de celdas fuera de límites para frecuencia lluvias: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/frecuencia_lluvias_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener opciones de frecuencia lluvias: {str(e)}"}), 500

# --- Función de Normalización de Texto ---
def normalizar_texto(texto):
    if texto is None:
        return ""
    texto = texto.lower().strip()
    # Eliminar prefijos comunes como "partido de "
    if texto.startswith('partido de '):
        texto = texto[len('partido de '):]

    # Reemplazos para vocales acentuadas y diéresis comunes en español
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ä': 'a', 'ë': 'e', 'ï': 'i', 'ö': 'o', 'ü': 'u', # Considerar si 'ü' debe ser 'u' o 'ue'
        'à': 'a', '': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u'  # Acentos graves, menos comunes en español pero por si acaso
        # Podrían añadirse más reemplazos si es necesario (ej. ñ -> n, aunque esto es debatible)
    }
    for acentuada, sin_acento in reemplazos.items():
        texto = texto.replace(acentuada, sin_acento)
    return texto

# --- NUEVA RUTA: Para buscar ciudad y obtener código (Refactorizada) ---
@app.route('/api/buscar_ciudad', methods=['POST'])
def buscar_ciudad():
    data = request.json
    full_address = data.get('full_address')

    if not full_address:
        return jsonify({"error": "Dirección completa no proporcionada."}), 400

    try:
        print(f"DEBUG: Solicitud a /api/buscar_ciudad para la dirección: '{full_address}'")

        df_ciudades = pd.read_excel(
            EXCEL_FILE_PATH,
            sheet_name='Ciudades',
            usecols="A,B",
            header=None,
            skiprows=1,
            nrows=1989,
            names=['codigo', 'ciudad_nombre'],
            engine='openpyxl'
        )

        address_parts = [part.strip() for part in full_address.split(',')]

        # Iterar a través de cada ciudad en nuestra lista de Excel
        for _, ciudad_row in df_ciudades.iterrows():
            codigo_ciudad = ciudad_row['codigo']
            nombre_ciudad_excel = str(ciudad_row['ciudad_nombre'])
            nombre_ciudad_excel_normalizado = normalizar_texto(nombre_ciudad_excel)

            if not nombre_ciudad_excel_normalizado:
                continue

            # Comparar cada parte de la direccin con la ciudad actual de la lista
            for part in address_parts:
                # Limpiar y normalizar la parte de la dirección
                part_cleaned = part.lower().strip()
                if part_cleaned.startswith("partido de "):
                    part_cleaned = part_cleaned.replace("partido de ", "", 1).strip()

                part_normalizada = normalizar_texto(part_cleaned)
                if not part_normalizada:
                    continue

                # Si la parte de la dirección coincide con una ciudad de la lista, la hemos encontrado
                if part_normalizada == nombre_ciudad_excel_normalizado:
                    print(f"DEBUG: Coincidencia encontrada. Parte de la dirección: '{part}' -> Ciudad en Excel: '{nombre_ciudad_excel}'. Código: {codigo_ciudad}")
                    return jsonify({"codigo_ciudad": codigo_ciudad, "nombre_ciudad": nombre_ciudad_excel, "message": f"Ciudad encontrada: {nombre_ciudad_excel}"}), 200

        print(f"WARN: No se encontró ninguna ciudad coincidente para la dirección: '{full_address}'")
        return jsonify({"codigo_ciudad": None, "message": f"No se pudo encontrar una ciudad válida en la dirección proporcionada."}), 200

    except FileNotFoundError:
        print(f"ERROR en /api/buscar_ciudad: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 500
    except KeyError as e:
        print(f"ERROR en /api/buscar_ciudad: Hoja 'Ciudades' o columnas A/B no encontradas? Error: {e}")
        return jsonify({"error": f"Error de clave al leer la hoja de cálculo para ciudades: {e}"}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/buscar_ciudad: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al buscar ciudad: {str(e)}"}), 500

# --- La ruta /api/escribir_dato_excel se ha eliminado para prevenir la corrupción de archivos.
# --- La lógica ahora maneja los datos en memoria en lugar de escribir en el archivo Excel. ---

# --- NUEVA RUTA: Para obtener el modelo de panel desde la celda C83 ---
@app.route('/api/get_modelo_panel', methods=['GET'])
def get_modelo_panel():
    """
    Lee y devuelve el valor de la celda C83 de la hoja 'Datos de Entrada'.
    """
    try:
        print(f"DEBUG: Solicitud a /api/get_modelo_panel. Leyendo celda C83 de 'Datos de Entrada' desde: {EXCEL_FILE_PATH}")

        # Usar openpyxl es más eficiente para una sola celda.
        import openpyxl

        # data_only=True para obtener el valor calculado si es una fórmula
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
        sheet = workbook['Datos de Entrada']
        valor_celda = sheet['C83'].value

        print(f"DEBUG: Valor leído de C83: {valor_celda}")

        return jsonify({"valor": valor_celda if valor_celda is not None else ""})

    except FileNotFoundError:
        print(f"ERROR en /api/get_modelo_panel: Archivo Excel no encontrado: {EXCEL_FILE_PATH}")
        return jsonify({"error": "Archivo Excel de configuración no encontrado."}), 404
    except KeyError:
        # Esto podría ocurrir si 'Datos de Entrada' no existe.
        print(f"ERROR en /api/get_modelo_panel: Hoja 'Datos de Entrada' no encontrada.")
        return jsonify({"error": "Hoja 'Datos de Entrada' no encontrada en el archivo Excel."}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/get_modelo_panel: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al obtener el modelo de panel: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener el modelo de panel dinámicamente ---
@app.route('/api/get_panel_model', methods=['POST'])
def get_panel_model_api():
    """
    Obtiene el nombre de un modelo de panel basado en la marca y potencia.
    """
    data = request.json
    marca = data.get('marca')
    potencia = data.get('potencia')

    if not marca or not potencia:
        return jsonify({"error": "Se requiere 'marca' y 'potencia'."}), 400

    try:
        print(f"DEBUG: API call to /api/get_panel_model. Marca: {marca}, Potencia: {potencia}")
        model_name = engine.get_panel_model_name(marca, potencia, EXCEL_FILE_PATH)

        if isinstance(model_name, dict) and "error" in model_name:
            return jsonify(model_name), 500

        return jsonify({"model_name": model_name})

    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/get_panel_model: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al buscar modelo de panel: {str(e)}"}), 500


# --- NUEVA RUTA: Para obtener la lista de modelos de inversores ---
@app.route('/api/get_inverter_options', methods=['GET'])
def get_inverter_options():
    """
    Lee el archivo 'inversores_genericos.json' y devuelve una lista de todos los
    inversores disponibles para poblar un dropdown en el frontend.
    """
    try:
        inversores_json_path = os.path.join(SCRIPT_DIR, 'data', 'inversores_genericos.json')
        with open(inversores_json_path, 'r', encoding='utf-8') as f:
            inversores_data = json.load(f)

        # Limpiar los datos de valores NaN antes de enviarlos
        cleaned_data = clean_nan_in_data(inversores_data)

        # Extraer solo los campos necesarios para el dropdown
        inverter_list = [
            {"NOMBRE": inv.get("NOMBRE"), "Pot nom CA [W]": inv.get("Pot nom CA [W]")}
            for inv in cleaned_data
        ]
        return jsonify(inverter_list)

    except FileNotFoundError:
        return jsonify({"error": "Archivo de datos de inversores no encontrado."}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/get_inverter_options: {e}")
        print(traceback.format_exc())
        return jsonify({"error": "Error interno del servidor al obtener opciones de inversor."}), 500


# --- CORREGIDO: Ruta para obtener una lista de inversores ---
# Esta ruta ahora simplemente devuelve todos los inversores del JSON.
# La lógica de filtrado se puede aadir en el futuro si es necesario,
# pero por ahora, esto soluciona el crash.
@app.route('/api/get_suitable_inverters', methods=['POST'])
def get_suitable_inverters_api():
    """
    Devuelve una lista completa de modelos de inversores desde el archivo JSON,
    asegurndose de que sea un JSON válido (sin NaN).
    """
    try:
        print("DEBUG: API call to /api/get_suitable_inverters. Returning full list from JSON.")
        inversores_json_path = os.path.join(SCRIPT_DIR, 'data', 'inversores_genericos.json')
        with open(inversores_json_path, 'r', encoding='utf-8') as f:
            inversores_data = json.load(f)

        # Limpiar los datos de valores NaN antes de enviarlos
        cleaned_data = clean_nan_in_data(inversores_data)

        # Devolver la lista completa y limpia.
        return jsonify(cleaned_data)

    except FileNotFoundError:
        print(f"ERROR en /api/get_suitable_inverters: No se encontró {inversores_json_path}")
        return jsonify({"error": "El archivo de datos de inversores no se encuentra en el servidor."}), 500
    except Exception as e:
        import traceback
        print(f"ERROR GENERAL en /api/get_suitable_inverters: {e}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor al buscar inversores: {str(e)}"}), 500

# --- NUEVAS RUTAS PARA UBICACIÓN E INFORME ---

from flask import Response  # asegurate de que esté importado arriba (por si no está)




# --- El endpoint temporal /api/verificar_celda ha sido eliminado. ---

# ======== AJUSTES BÁSICOS Y HELPERS ========
import os
from flask import jsonify, request
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(
    BASE_DIR,
    "Calculador Solar - web 06-24_con ayuda - modificaciones 2025_5.xlsx"   # <-- usa el nombre exacto de TU archivo bueno
)

def _open_wb():
    """Abrir workbook sin evaluar fórmulas (rápido) y con logs."""
    if not os.path.exists(EXCEL_FILE_PATH):
        raise FileNotFoundError(f"No existe: {EXCEL_FILE_PATH}")
    wb = load_workbook(
    EXCEL_FILE_PATH,
    data_only=False,
    read_only=False,
    keep_vba=False,
    keep_links=False   # evita parseo de vínculos externos que puede colgar
)
    return wb

# ======== ENDPOINTS DE DIAGNÓSTICO ========

@app.route('/api/ping')
def ping():
    return jsonify({'pong': True})

@app.route('/api/health')
def health():
    return jsonify({
        'ok': True,
        'cwd': os.getcwd(),
        'excel_path': EXCEL_FILE_PATH,
        'excel_exists': os.path.exists(EXCEL_FILE_PATH),
    })

# ======== ENDPOINTS REALES ========

@app.route('/api/ubicacion', methods=['POST'])
def set_ubicacion():
    """
    Guarda la ciudad en B7 de 'Datos de Entrada'.
    Protegido con lock, logs detallados y guardado a archivo temporal.
    """
    import time
    import json
    import shutil
    from tempfile import NamedTemporaryFile

    t0 = time.time()
    try:
        print("[/api/ubicacion] INICIO", flush=True)
        payload = request.get_json(force=True) or {}
        city = (payload.get('city') or '').strip()
        lat = payload.get('lat'); lng = payload.get('lng')
        print(f"[/api/ubicacion] payload={payload}", flush=True)

        if not city:
            print("[/api/ubicacion] city vacío", flush=True)
            return jsonify({'error': 'city es requerido'}), 400

        if not os.path.exists(EXCEL_FILE_PATH):
            print(f"[/api/ubicacion] NO existe Excel: {EXCEL_FILE_PATH}", flush=True)
            return jsonify({'error': 'Archivo Excel no encontrado.'}), 404

        # BLOQUEO EXCLUSIVO para evitar colgues por acceso concurrente
        print("[/api/ubicacion] esperando excel_lock...", flush=True)
        with excel_lock:
            print("[/api/ubicacion] excel_lock OBTENIDO", flush=True)

            wb = load_workbook(
              EXCEL_FILE_PATH,
              data_only=False,
              read_only=False,
              keep_vba=False,
              keep_links=False
            )
            print(f"[/api/ubicacion] hojas={wb.sheetnames}", flush=True)
            try:
                print("[/api/ubicacion] load_workbook...", flush=True)
                wb = load_workbook(EXCEL_FILE_PATH, data_only=False, read_only=False, keep_vba=False)
                print(f"[/api/ubicacion] hojas={wb.sheetnames}", flush=True)

                sheet_name = 'Datos de Entrada'  # AJUSTAR si tu hoja tiene otro nombre exacto
                if sheet_name not in wb.sheetnames:
                    raise KeyError(f"Hoja no encontrada: '{sheet_name}' (disponibles: {wb.sheetnames})")

                ws = wb[sheet_name]
                ws['B7'] = city
                print(f"[/api/ubicacion] B7 <- {city}", flush=True)

                # Guardado a archivo temporal + reemplazo atómico
                tmpf = NamedTemporaryFile(delete=False, dir=os.path.dirname(EXCEL_FILE_PATH), suffix='.xlsx')
                tmpf_path = tmpf.name
                tmpf.close()

                print(f"[/api/ubicacion] wb.save({tmpf_path}) ...", flush=True)
                wb.save(tmpf_path)
                print("[/api/ubicacion] wb.save OK", flush=True)

                print(f"[/api/ubicacion] reemplazo atómico -> {EXCEL_FILE_PATH}", flush=True)
                shutil.move(tmpf_path, EXCEL_FILE_PATH)
                print("[/api/ubicacion] move OK", flush=True)

            finally:
                if wb is not None:
                    wb.close()
                    print("[/api/ubicacion] wb.close()", flush=True)

        t1 = time.time()
        print(f"[/api/ubicacion] FIN ok en {t1 - t0:.2f}s", flush=True)
        return jsonify({'ok': True, 'city': city, 'lat': lat, 'lng': lng})

    except Exception as e:
        import traceback
        print("[/api/ubicacion] ERROR:", e, flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500



# --- INFORME: lee Resultados!B3:L50 ---
@app.route('/api/informe', methods=['GET'], endpoint='api_informe')
def api_informe():
    try:
        import pandas as pd
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({'error': f'Excel no encontrado: {EXCEL_FILE_PATH}'}), 404

        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Resultados', engine='openpyxl', header=None)
        # B3:L50 -> filas 2..49 (0-based), columnas 1..11
        resultados = df.iloc[2:50, 1:12].fillna('').values.tolist()
        print(f"[/api/informe] filas={len(resultados)}", flush=True)
        return jsonify({'ok': True, 'resultados': resultados})

    except Exception as e:
        import traceback
        print("[/api/informe] ERROR:", e, flush=True)
        print(traceback.format_exc(), flush=True)
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
